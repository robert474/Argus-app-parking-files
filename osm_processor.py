#!/usr/bin/env python3
"""
OSM Truck Parking Processor
Converts Overpass API exports (GeoJSON/JSON) to clean CSV/Excel for Argus AI.

Usage:
    python osm_processor.py rest_areas.geojson services.geojson truck_parking.geojson

Output:
    osm_truck_parking_combined.csv
    osm_truck_parking_combined.xlsx
"""

import json
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Note: openpyxl not installed. CSV will be created but not Excel.")
    print("Install with: pip install openpyxl")


# Fields to extract from OSM tags
FIELDS = [
    'osm_id',
    'osm_type',
    'name',
    'latitude',
    'longitude',
    'highway',
    'amenity',
    'hgv',
    'capacity_hgv',
    'capacity',
    'toilets',
    'drinking_water',
    'shelter',
    'picnic_table',
    'dump_station',
    'operator',
    'ref',
    'addr_state',
    'addr_city',
    'source_file'
]


def parse_overpass_json(filepath: str) -> list:
    """Parse Overpass API JSON/GeoJSON export."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    records = []
    source_name = Path(filepath).stem

    # Handle GeoJSON format (from Overpass Turbo export)
    if data.get('type') == 'FeatureCollection':
        for feature in data.get('features', []):
            record = extract_geojson_feature(feature, source_name)
            if record:
                records.append(record)

    # Handle raw Overpass JSON format
    elif 'elements' in data:
        for element in data.get('elements', []):
            record = extract_overpass_element(element, source_name)
            if record:
                records.append(record)

    return records


def extract_geojson_feature(feature: dict, source_name: str) -> dict:
    """Extract record from GeoJSON feature."""
    props = feature.get('properties', {})
    geom = feature.get('geometry', {})

    # Get coordinates
    coords = geom.get('coordinates', [])
    if geom.get('type') == 'Point' and len(coords) >= 2:
        lon, lat = coords[0], coords[1]
    elif geom.get('type') == 'Polygon' and coords:
        # Use centroid for polygons (ways)
        all_coords = coords[0] if coords else []
        if all_coords:
            lon = sum(c[0] for c in all_coords) / len(all_coords)
            lat = sum(c[1] for c in all_coords) / len(all_coords)
        else:
            lon, lat = None, None
    else:
        lon, lat = None, None

    return {
        'osm_id': props.get('id') or props.get('@id', ''),
        'osm_type': props.get('type') or props.get('@type', 'node'),
        'name': props.get('name', ''),
        'latitude': lat,
        'longitude': lon,
        'highway': props.get('highway', ''),
        'amenity': props.get('amenity', ''),
        'hgv': props.get('hgv', ''),
        'capacity_hgv': props.get('capacity:hgv', ''),
        'capacity': props.get('capacity', ''),
        'toilets': props.get('toilets', ''),
        'drinking_water': props.get('drinking_water', ''),
        'shelter': props.get('shelter', ''),
        'picnic_table': props.get('picnic_table', ''),
        'dump_station': props.get('dump_station', ''),
        'operator': props.get('operator', ''),
        'ref': props.get('ref', ''),
        'addr_state': props.get('addr:state', ''),
        'addr_city': props.get('addr:city', ''),
        'source_file': source_name
    }


def extract_overpass_element(element: dict, source_name: str) -> dict:
    """Extract record from raw Overpass element."""
    tags = element.get('tags', {})

    # Get coordinates
    if element.get('type') == 'node':
        lat = element.get('lat')
        lon = element.get('lon')
    elif element.get('type') == 'way' and 'center' in element:
        lat = element['center'].get('lat')
        lon = element['center'].get('lon')
    else:
        lat, lon = None, None

    return {
        'osm_id': element.get('id', ''),
        'osm_type': element.get('type', 'node'),
        'name': tags.get('name', ''),
        'latitude': lat,
        'longitude': lon,
        'highway': tags.get('highway', ''),
        'amenity': tags.get('amenity', ''),
        'hgv': tags.get('hgv', ''),
        'capacity_hgv': tags.get('capacity:hgv', ''),
        'capacity': tags.get('capacity', ''),
        'toilets': tags.get('toilets', ''),
        'drinking_water': tags.get('drinking_water', ''),
        'shelter': tags.get('shelter', ''),
        'picnic_table': tags.get('picnic_table', ''),
        'dump_station': tags.get('dump_station', ''),
        'operator': tags.get('operator', ''),
        'ref': tags.get('ref', ''),
        'addr_state': tags.get('addr:state', ''),
        'addr_city': tags.get('addr:city', ''),
        'source_file': source_name
    }


def deduplicate_records(records: list) -> list:
    """Remove duplicate OSM IDs, keeping first occurrence."""
    seen = set()
    unique = []
    for record in records:
        key = (record['osm_id'], record['osm_type'])
        if key not in seen:
            seen.add(key)
            unique.append(record)
    return unique


def write_csv(records: list, output_path: str):
    """Write records to CSV."""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(records)
    print(f"Created: {output_path} ({len(records)} records)")


def write_excel(records: list, output_path: str):
    """Write records to formatted Excel file."""
    if not HAS_OPENPYXL:
        print("Skipping Excel output (openpyxl not installed)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OSM Truck Parking"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, field in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Write data
    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(field, ''))
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, field in enumerate(FIELDS, 1):
        max_length = len(field)
        for row in range(2, min(len(records) + 2, 100)):  # Sample first 100 rows
            cell_value = str(ws.cell(row=row, column=col_idx).value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add metadata sheet
    meta = wb.create_sheet(title="Metadata")
    meta['A1'] = "Generated"
    meta['B1'] = datetime.now().isoformat()
    meta['A2'] = "Total Records"
    meta['B2'] = len(records)
    meta['A3'] = "Source"
    meta['B3'] = "OpenStreetMap via Overpass API"

    wb.save(output_path)
    print(f"Created: {output_path} ({len(records)} records)")


def main():
    if len(sys.argv) < 2:
        print("Usage: python osm_processor.py <file1.geojson> [file2.geojson] ...")
        print("\nExample:")
        print("  python osm_processor.py rest_areas.geojson services.geojson truck_parking.geojson")
        sys.exit(1)

    all_records = []

    for filepath in sys.argv[1:]:
        if not Path(filepath).exists():
            print(f"Warning: File not found: {filepath}")
            continue

        print(f"Processing: {filepath}")
        records = parse_overpass_json(filepath)
        print(f"  Found {len(records)} records")
        all_records.extend(records)

    if not all_records:
        print("No records found. Check input files.")
        sys.exit(1)

    # Deduplicate
    unique_records = deduplicate_records(all_records)
    print(f"\nTotal unique records: {len(unique_records)}")

    # Write outputs
    write_csv(unique_records, 'osm_truck_parking_combined.csv')
    write_excel(unique_records, 'osm_truck_parking_combined.xlsx')

    # Summary by type
    print("\nBreakdown by type:")
    by_highway = {}
    by_amenity = {}
    for r in unique_records:
        if r['highway']:
            by_highway[r['highway']] = by_highway.get(r['highway'], 0) + 1
        if r['amenity']:
            by_amenity[r['amenity']] = by_amenity.get(r['amenity'], 0) + 1

    for k, v in sorted(by_highway.items(), key=lambda x: -x[1]):
        print(f"  highway={k}: {v}")
    for k, v in sorted(by_amenity.items(), key=lambda x: -x[1]):
        print(f"  amenity={k}: {v}")


if __name__ == '__main__':
    main()
